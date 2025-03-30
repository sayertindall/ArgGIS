#!/usr/bin/env python3
"""
CSV Utilities Module

A utility module for handling CSV and tabular data operations.
"""

import pandas as pd
import numpy as np
import os
import logging
from pathlib import Path
from typing import List, Dict, Union, Optional, Any, Tuple

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler()
    ]
)

logger = logging.getLogger("CSVUtils")

def load_csv(path: Union[str, Path], encoding='utf-8-sig') -> pd.DataFrame:
    """Load a CSV file into a pandas DataFrame.
    
    Args:
        path: Path to the CSV file
        encoding: File encoding (default: utf-8-sig to handle BOM)
        
    Returns:
        DataFrame: Loaded CSV data
    """
    try:
        logger.info(f"Loading CSV from {path}")
        df = pd.read_csv(path, encoding=encoding)
        logger.info(f"Loaded CSV with {len(df)} rows and {len(df.columns)} columns")
        return df
    except Exception as e:
        logger.error(f"Error loading CSV: {e}")
        return None

def preview(df: pd.DataFrame, n: int = 3) -> pd.DataFrame:
    """Preview the first n rows of a DataFrame.
    
    Args:
        df: DataFrame to preview
        n: Number of rows to preview
        
    Returns:
        DataFrame: First n rows of the DataFrame
    """
    if df is None or len(df) == 0:
        logger.warning("Cannot preview: DataFrame is empty or None")
        return None
        
    return df.head(n)

def validate_columns(df: pd.DataFrame, required: List[str]) -> bool:
    """Validate that a DataFrame contains required columns.
    
    Args:
        df: DataFrame to validate
        required: List of required column names
        
    Returns:
        bool: Whether the DataFrame contains all required columns
    """
    if df is None:
        logger.warning("Cannot validate columns: DataFrame is None")
        return False
        
    missing = [col for col in required if col not in df.columns]
    if missing:
        logger.warning(f"Missing required columns: {missing}")
        return False
        
    logger.info(f"All required columns present: {required}")
    return True

def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Clean column names by removing spaces, special characters, and standardizing case.
    
    Args:
        df: DataFrame to clean
        
    Returns:
        DataFrame: DataFrame with cleaned column names
    """
    if df is None:
        logger.warning("Cannot clean column names: DataFrame is None")
        return None
        
    logger.info("Cleaning column names")
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_').str.replace('[^\w_]', '', regex=True)
    return df

def drop_empty_columns(df: pd.DataFrame, threshold: float = 0.9) -> pd.DataFrame:
    """Drop columns that have more than threshold% missing values.
    
    Args:
        df: DataFrame to process
        threshold: Threshold for dropping columns (default: 0.9 means drop if 90% or more values are missing)
        
    Returns:
        DataFrame: DataFrame with empty columns dropped
    """
    if df is None:
        logger.warning("Cannot drop empty columns: DataFrame is None")
        return None
        
    logger.info(f"Dropping columns with more than {threshold*100}% missing values")
    before_cols = len(df.columns)
    df = df.copy()
    df = df.dropna(axis=1, thresh=int(len(df) * (1 - threshold)))
    after_cols = len(df.columns)
    
    if before_cols > after_cols:
        logger.info(f"Dropped {before_cols - after_cols} columns")
    else:
        logger.info("No columns were dropped")
        
    return df

def find_duplicate_rows(df: pd.DataFrame, subset: Optional[List[str]] = None) -> pd.DataFrame:
    """Find duplicate rows in a DataFrame.
    
    Args:
        df: DataFrame to check for duplicates
        subset: List of columns to consider for duplicates (default: all columns)
        
    Returns:
        DataFrame: Duplicate rows
    """
    if df is None:
        logger.warning("Cannot find duplicates: DataFrame is None")
        return None
        
    logger.info(f"Finding duplicate rows{' based on subset: ' + str(subset) if subset else ''}")
    duplicates = df[df.duplicated(subset=subset, keep='first')]
    logger.info(f"Found {len(duplicates)} duplicate rows")
    return duplicates

def export_to_csv(df: pd.DataFrame, output_path: str, index: bool = False) -> bool:
    """Export a DataFrame to a CSV file.
    
    Args:
        df: DataFrame to export
        output_path: Path for the output file
        index: Whether to include the index in the output
        
    Returns:
        bool: Whether the export was successful
    """
    if df is None:
        logger.warning("Cannot export: DataFrame is None")
        return False
        
    try:
        logger.info(f"Exporting DataFrame to {output_path}")
        df.to_csv(output_path, index=index)
        logger.info(f"Successfully exported {len(df)} rows to {output_path}")
        return True
    except Exception as e:
        logger.error(f"Error exporting to CSV: {e}")
        return False

def summarize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Generate a summary of columns in a DataFrame.
    
    Args:
        df: DataFrame to summarize
        
    Returns:
        DataFrame: Summary of columns including data type, unique values, missing values, etc.
    """
    if df is None:
        logger.warning("Cannot summarize columns: DataFrame is None")
        return None
        
    logger.info("Generating column summary")
    summary = []
    for col in df.columns:
        dtype = df[col].dtype
        unique_count = df[col].nunique()
        non_null_count = df[col].count()
        null_percentage = (len(df) - non_null_count) / len(df) * 100
        
        summary.append({
            'column': col,
            'dtype': str(dtype),
            'unique_values': unique_count,
            'null_percentage': null_percentage,
            'sample_values': str(df[col].dropna().sample(min(3, non_null_count)).tolist())
        })
    
    return pd.DataFrame(summary)

def merge_csv_files(file_paths: List[str], on: Union[str, List[str]], how: str = 'inner') -> pd.DataFrame:
    """Merge multiple CSV files into a single DataFrame.
    
    Args:
        file_paths: List of paths to CSV files
        on: Column(s) to join on
        how: Type of merge to perform ('inner', 'outer', 'left', 'right')
        
    Returns:
        DataFrame: Merged data
    """
    if not file_paths:
        logger.warning("Cannot merge: No file paths provided")
        return None
        
    logger.info(f"Merging {len(file_paths)} CSV files")
    dfs = []
    for path in file_paths:
        df = load_csv(path)
        if df is not None:
            dfs.append(df)
    
    if not dfs:
        logger.warning("No valid DataFrames to merge")
        return None
        
    result = dfs[0]
    for i, df in enumerate(dfs[1:], 1):
        logger.info(f"Merging DataFrame {i+1}")
        result = pd.merge(result, df, on=on, how=how)
    
    logger.info(f"Merged result has {len(result)} rows and {len(result.columns)} columns")
    return result

def filter_by_value(df: pd.DataFrame, column: str, value: Any, operator: str = '==') -> pd.DataFrame:
    """Filter a DataFrame by a value in a column.
    
    Args:
        df: DataFrame to filter
        column: Column to filter on
        value: Value to filter by
        operator: Comparison operator ('==', '!=', '>', '<', '>=', '<=')
        
    Returns:
        DataFrame: Filtered DataFrame
    """
    if df is None:
        logger.warning("Cannot filter: DataFrame is None")
        return None
        
    if column not in df.columns:
        logger.warning(f"Column '{column}' not found in DataFrame")
        return df
        
    logger.info(f"Filtering by {column} {operator} {value}")
    
    if operator == '==':
        mask = df[column] == value
    elif operator == '!=':
        mask = df[column] != value
    elif operator == '>':
        mask = df[column] > value
    elif operator == '<':
        mask = df[column] < value
    elif operator == '>=':
        mask = df[column] >= value
    elif operator == '<=':
        mask = df[column] <= value
    else:
        logger.error(f"Unsupported operator: {operator}")
        return df
        
    result = df[mask]
    logger.info(f"Filter returned {len(result)} rows")
    return result

def pivot_table(df: pd.DataFrame, index: Union[str, List[str]], columns: str, values: str, 
               aggfunc: str = 'mean') -> pd.DataFrame:
    """Create a pivot table from a DataFrame.
    
    Args:
        df: DataFrame to pivot
        index: Column(s) to use as index
        columns: Column to use as columns
        values: Column to use as values
        aggfunc: Aggregation function ('mean', 'sum', 'count', etc.)
        
    Returns:
        DataFrame: Pivot table
    """
    if df is None:
        logger.warning("Cannot create pivot table: DataFrame is None")
        return None
        
    logger.info(f"Creating pivot table with index={index}, columns={columns}, values={values}")
    try:
        pivot = pd.pivot_table(df, index=index, columns=columns, values=values, aggfunc=aggfunc)
        logger.info(f"Created pivot table with shape {pivot.shape}")
        return pivot
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        return None

def detect_outliers(df: pd.DataFrame, column: str, method: str = 'iqr', threshold: float = 1.5) -> pd.DataFrame:
    """Detect outliers in a numeric column.
    
    Args:
        df: DataFrame to analyze
        column: Numeric column to check for outliers
        method: Method to use ('iqr' or 'zscore')
        threshold: Threshold for outlier detection
        
    Returns:
        DataFrame: Rows containing outliers
    """
    if df is None:
        logger.warning("Cannot detect outliers: DataFrame is None")
        return None
        
    if column not in df.columns:
        logger.warning(f"Column '{column}' not found in DataFrame")
        return None
        
    if not np.issubdtype(df[column].dtype, np.number):
        logger.warning(f"Column '{column}' is not numeric")
        return None
        
    logger.info(f"Detecting outliers in column '{column}' using {method} method")
    
    if method == 'iqr':
        Q1 = df[column].quantile(0.25)
        Q3 = df[column].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - threshold * IQR
        upper_bound = Q3 + threshold * IQR
        outliers = df[(df[column] < lower_bound) | (df[column] > upper_bound)]
    elif method == 'zscore':
        from scipy import stats
        z_scores = np.abs(stats.zscore(df[column]))
        outliers = df[z_scores > threshold]
    else:
        logger.error(f"Unsupported method: {method}")
        return None
        
    logger.info(f"Found {len(outliers)} outliers")
    return outliers

def find_missing_values(df: pd.DataFrame) -> pd.DataFrame:
    """Find and summarize missing values in a DataFrame.
    
    Args:
        df: DataFrame to analyze
        
    Returns:
        DataFrame: Summary of missing values by column
    """
    if df is None:
        logger.warning("Cannot find missing values: DataFrame is None")
        return None
        
    logger.info("Analyzing missing values")
    missing = pd.DataFrame({
        'column': df.columns,
        'missing_count': df.isnull().sum().values,
        'missing_percentage': (df.isnull().sum() / len(df) * 100).values
    })
    
    missing = missing.sort_values('missing_count', ascending=False)
    logger.info(f"Found columns with missing values: {sum(missing['missing_count'] > 0)}")
    return missing

def analyze_excel_file(file_path):
    """Analyzes Excel file structure and content with detailed report."""
    import pandas as pd
    import openpyxl
    
    # Basic file info
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    
    results = {
        "file_name": file_path,
        "sheet_count": len(sheets),
        "sheets": {}
    }
    
    # Analyze each sheet
    for sheet in sheets:
        df = pd.read_excel(file_path, sheet_name=sheet)
        results["sheets"][sheet] = {
            "row_count": len(df),
            "column_count": len(df.columns),
            "columns": list(df.columns),
            "missing_values": df.isna().sum().to_dict(),
            "duplicated_rows": df.duplicated().sum(),
            "data_types": df.dtypes.astype(str).to_dict()
        }
    
    return results

def smart_type_converter(df):
    """Intelligently converts column types based on content analysis."""
    import pandas as pd
    import numpy as np
    
    result_df = df.copy()
    
    for col in result_df.columns:
        # Skip columns that are already non-object type
        if result_df[col].dtype != 'object':
            continue
            
        # Try to convert to datetime
        try:
            result_df[col] = pd.to_datetime(result_df[col], errors='raise')
            continue
        except:
            pass
            
        # Try to convert to numeric
        if result_df[col].str.match(r'^-?\d*\.?\d+$').all():
            try:
                result_df[col] = pd.to_numeric(result_df[col])
                continue
            except:
                pass
                
        # Clean strings
        result_df[col] = result_df[col].str.strip()
        
    return result_df

def standardize_excel(input_path, output_path=None):
    """Standardizes Excel files with consistent formatting and structure."""
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Read all sheets
    excel_file = pd.ExcelFile(input_path)
    all_dfs = {sheet: pd.read_excel(excel_file, sheet_name=sheet) for sheet in excel_file.sheet_names}
    
    # Standardize each dataframe
    for sheet, df in all_dfs.items():
        # Standardize column names
        df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]
        
        # Remove duplicate rows
        df = df.drop_duplicates()
        
        # Fill small gaps in data
        df = df.fillna(method='ffill', limit=3)
        
        all_dfs[sheet] = df
    
    # Write back with formatting
    if output_path is None:
        output_path = input_path.replace('.xlsx', '_standardized.xlsx')
        
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet, df in all_dfs.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            
            # Apply formatting
            ws = writer.sheets[sheet]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                
    return output_path

def deduplicate_with_fuzzy(df, text_columns, threshold=80):
    """Removes duplicates using fuzzy string matching."""
    from fuzzywuzzy import fuzz
    import pandas as pd
    import numpy as np
    
    # Generate fingerprints from text columns
    df['_fingerprint'] = df[text_columns].astype(str).agg(' '.join, axis=1)
    
    # Track duplicates
    duplicate_indices = []
    processed_indices = []
    
    for i, row in enumerate(df['_fingerprint']):
        if i in processed_indices:
            continue
            
        processed_indices.append(i)
        current_duplicates = []
        
        # Compare with remaining rows
        for j in range(i+1, len(df)):
            if j in processed_indices:
                continue
                
            similarity = fuzz.ratio(row, df['_fingerprint'].iloc[j])
            if similarity >= threshold:
                current_duplicates.append(j)
                processed_indices.append(j)
                
        duplicate_indices.extend(current_duplicates)
    
    # Remove duplicates
    clean_df = df.drop(duplicate_indices).drop('_fingerprint', axis=1)
    return clean_df

def auto_map_columns(source_df, target_template_df):
    """Automatically maps columns between dataframes based on similarity."""
    from fuzzywuzzy import fuzz
    
    source_cols = list(source_df.columns)
    target_cols = list(target_template_df.columns)
    mapping = {}
    
    for target_col in target_cols:
        # Find best match in source columns
        best_match = None
        best_score = 0
        
        for source_col in source_cols:
            score = fuzz.ratio(target_col.lower(), source_col.lower())
            if score > best_score and score > 60:  # Minimum threshold
                best_score = score
                best_match = source_col
                
        if best_match:
            mapping[target_col] = best_match
    
    # Create new dataframe with mapped columns
    result_df = pd.DataFrame()
    for target_col, source_col in mapping.items():
        result_df[target_col] = source_df[source_col]
        
    # Report unmapped columns
    unmapped_target = set(target_cols) - set(mapping.keys())
    
    return result_df, mapping, unmapped_target

def hierarchical_normalizer(df, hierarchy_cols, value_cols):
    """
    Normalizes hierarchical data with parent-child relationships.
    Handles multi-level data structures common in financial/org reports.
    """
    import pandas as pd
    import numpy as np
    
    result = df.copy()
    
    # Create hierarchical index
    result.set_index(hierarchy_cols, inplace=True)
    
    # Identify leaf nodes vs aggregation nodes
    is_leaf = ~result.index.duplicated(keep=False)
    
    # Normalize values within each hierarchy
    for col in value_cols:
        # Get sums at each level
        level_sums = {}
        for i in range(len(hierarchy_cols)):
            level_index = result.index.droplevel(list(range(i+1, len(hierarchy_cols))))
            level_sums[i] = result.groupby(level=level_index)[col].transform('sum')
        
        # Calculate proportions at each level
        for i in range(len(hierarchy_cols)):
            level_prop_col = f"{col}_prop_L{i+1}"
            with np.errstate(divide='ignore', invalid='ignore'):
                result[level_prop_col] = result[col] / level_sums[i]
                result[level_prop_col].fillna(0, inplace=True)
    
    return result.reset_index()

def distribution_matcher(df, column, target_distribution):
    """
    Transforms values in a column to match a target statistical distribution.
    Preserves rank order while normalizing to match target parameters.
    """
    import pandas as pd
    import numpy as np
    from scipy import stats
    
    # Extract values and ranks
    values = df[column].values
    ranks = stats.rankdata(values, method='average')
    n = len(values)
    
    # Calculate percentiles
    percentiles = (ranks - 0.5) / n
    
    # Generate new values based on target distribution
    if target_distribution == 'normal':
        # Standard normal distribution
        new_values = stats.norm.ppf(percentiles)
    elif target_distribution == 'uniform':
        # Uniform distribution between 0 and 1
        new_values = percentiles
    elif target_distribution == 'exponential':
        # Exponential distribution
        new_values = stats.expon.ppf(percentiles)
    else:
        raise ValueError("Unsupported distribution")
    
    # Create result with transformed values
    result = df.copy()
    result[f"{column}_normalized"] = new_values
    
    return result

def auto_categorizer(df, column, method='jenks', num_categories=5):
    """
    Automatically categorizes continuous data using various statistical methods.
    Creates optimal breakpoints based on data distribution.
    """
    import pandas as pd
    import numpy as np
    
    values = df[column].dropna().values
    result = df.copy()
    
    if method == 'jenks':
        # Natural breaks (Jenks) algorithm
        from jenkspy import JenksNaturalBreaks
        
        # Find natural breakpoints
        jnb = JenksNaturalBreaks(num_categories)
        breaks = jnb.fit(values)
        
        # Create categories
        result[f"{column}_category"] = pd.cut(
            result[column], 
            bins=breaks, 
            include_lowest=True,
            labels=[f"Category {i+1}" for i in range(num_categories)]
        )
        
    elif method == 'quantile':
        # Quantile-based categorization
        quantiles = np.linspace(0, 1, num_categories+1)
        breaks = [np.min(values) - 0.001] + list(np.quantile(values, quantiles[1:-1])) + [np.max(values) + 0.001]
        
        result[f"{column}_category"] = pd.cut(
            result[column], 
            bins=breaks,
            labels=[f"Q{i+1}" for i in range(num_categories)]
        )
        
    elif method == 'kmeans':
        # K-means clustering
        from sklearn.cluster import KMeans
        
        # Reshape for sklearn
        X = values.reshape(-1, 1)
        
        # Apply k-means
        kmeans = KMeans(n_clusters=num_categories, random_state=0).fit(X)
        
        # Get cluster centers and sort them
        centers = sorted(kmeans.cluster_centers_.flatten())
        
        # Predict clusters for all values including NaN (assign NaN to its own category)
        result[f"{column}_category"] = None
        mask = ~df[column].isna()
        result.loc[mask, f"{column}_category"] = kmeans.predict(df.loc[mask, column].values.reshape(-1, 1))
        
    return result

def seasonal_normalizer(df, date_column, value_column, periods=None):
    """
    Normalizes time series data to account for seasonal patterns.
    Detects and removes seasonality for proper comparison across time periods.
    """
    import pandas as pd
    import numpy as np
    
    # Ensure datetime format
    if df[date_column].dtype != 'datetime64[ns]':
        df[date_column] = pd.to_datetime(df[date_column])
    
    result = df.copy()
    
    # Extract time components
    result['year'] = result[date_column].dt.year
    result['month'] = result[date_column].dt.month
    result['day'] = result[date_column].dt.day
    result['weekday'] = result[date_column].dt.weekday
    result['quarter'] = result[date_column].dt.quarter
    
    # Default periods to check
    if periods is None:
        periods = ['month', 'quarter', 'weekday']
    
    # Calculate seasonal indices for each period type
    for period in periods:
        # Group by period and calculate mean
        period_means = result.groupby(period)[value_column].mean()
        
        # Calculate overall mean
        overall_mean = result[value_column].mean()
        
        # Calculate seasonal indices
        seasonal_indices = period_means / overall_mean
        
        # Map indices back to original data
        result[f"{period}_index"] = result[period].map(seasonal_indices)
        
        # Create seasonally adjusted values
        result[f"{value_column}_adj_{period}"] = result[value_column] / result[f"{period}_index"]
    
    return result

def smart_outlier_normalizer(df, columns, method='iqr', replace_method='winsorize'):
    """
    Advanced outlier detection and normalization with context-aware replacements.
    Handles different data distributions appropriately.
    """
    import pandas as pd
    import numpy as np
    from scipy import stats
    
    result = df.copy()
    
    for col in columns:
        # Skip non-numeric columns
        if not pd.api.types.is_numeric_dtype(result[col]):
            continue
            
        values = result[col].dropna()
        
        # Check for distribution type
        skewness = stats.skew(values)
        is_skewed = abs(skewness) > 1.0
        
        outlier_mask = np.zeros(len(result), dtype=bool)
        
        if method == 'iqr':
            # IQR method
            Q1 = values.quantile(0.25)
            Q3 = values.quantile(0.75)
            IQR = Q3 - Q1
            
            # Adjust bounds based on distribution
            if is_skewed:
                # For skewed distributions, use asymmetric bounds
                if skewness > 0:  # Right-skewed
                    lower_bound = Q1 - 1.5 * IQR
                    upper_bound = Q3 + 3.0 * IQR
                else:  # Left-skewed
                    lower_bound = Q1 - 3.0 * IQR
                    upper_bound = Q3 + 1.5 * IQR
            else:
                # For normal-like distributions
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                
            # Create mask
            col_mask = (result[col] < lower_bound) | (result[col] > upper_bound)
            outlier_mask = outlier_mask | col_mask
            
            # Store original values
            result[f"{col}_original"] = result[col].copy()
            
            # Replace outliers
            if replace_method == 'winsorize':
                result.loc[result[col] < lower_bound, col] = lower_bound
                result.loc[result[col] > upper_bound, col] = upper_bound
            elif replace_method == 'mean':
                result.loc[outlier_mask, col] = values.mean()
            elif replace_method == 'median':
                result.loc[outlier_mask, col] = values.median()
            elif replace_method == 'nearest_valid':
                # Replace with nearest valid value
                for idx in result[outlier_mask].index:
                    if result.loc[idx, col] < lower_bound:
                        result.loc[idx, col] = lower_bound
                    else:
                        result.loc[idx, col] = upper_bound
    
    # Track which values were modified
    result['contains_normalized_outliers'] = outlier_mask
    
    return result

def cross_column_normalizer(df, related_columns, method='ratio'):
    """
    Normalizes values across related columns to ensure consistency.
    Handles scenarios where multiple columns should maintain certain relationships.
    """
    import pandas as pd
    import numpy as np
    
    result = df.copy()
    
    # Calculate summary statistics for normalization reference
    if method == 'ratio':
        # Sum across columns for each row
        row_sums = result[related_columns].sum(axis=1)
        
        # Create normalized columns as ratios of sum
        for col in related_columns:
            norm_col = f"{col}_norm"
            with np.errstate(divide='ignore', invalid='ignore'):
                result[norm_col] = result[col] / row_sums
                result[norm_col].fillna(0, inplace=True)
                
    elif method == 'zscore':
        # Z-score normalization within each column
        for col in related_columns:
            mean = result[col].mean()
            std = result[col].std()
            
            norm_col = f"{col}_zscore"
            with np.errstate(divide='ignore', invalid='ignore'):
                result[norm_col] = (result[col] - mean) / std
                result[norm_col].fillna(0, inplace=True)
                
    elif method == 'minmax':
        # Min-max scaling within each column
        for col in related_columns:
            min_val = result[col].min()
            max_val = result[col].max()
            
            norm_col = f"{col}_scaled"
            with np.errstate(divide='ignore', invalid='ignore'):
                result[norm_col] = (result[col] - min_val) / (max_val - min_val)
                result[norm_col].fillna(0, inplace=True)
    
    return result

def context_aware_text_normalizer(df, text_column, locale='es_ES'):
    """
    Advanced text normalization with locale-specific rules.
    Handles special cases for Spanish text including accents and formal/informal forms.
    """
    import pandas as pd
    import re
    import unicodedata
    
    result = df.copy()
    
    # Basic cleaning
    result[f"{text_column}_norm"] = result[text_column].astype(str).str.strip()
    
    # Locale-specific processing
    if locale.startswith('es'):  # Spanish
        # Handle accents (normalize without removing them)
        result[f"{text_column}_norm"] = result[f"{text_column}_norm"].apply(
            lambda x: unicodedata.normalize('NFC', x)
        )
        
        # Standardize common variations
        replacements = {
            r'\bUd\b|\bUd\.\b|\bUds\.\b|\bUsted\b': 'Usted',  # Formality
            r'\bDon\b|\bD\.\b': 'Don',  # Titles
            r'\bDoña\b|\bDña\.\b': 'Doña',
            r'\bSr\.\b|\bSeñor\b': 'Sr.',
            r'\bSra\.\b|\bSeñora\b': 'Sra.',
            r'\bAv\.\b|\bAvda\.\b|\bAvenida\b': 'Av.',  # Addresses
            r'\bC/\b|\bCalle\b': 'C/',
            r'\bPza\.\b|\bPlaza\b': 'Pza.',
            # Money and numbers
            r'€|\bEUR\b|\beuros?\b': '€',
            r'\bUS\$|\bUSD\b': 'USD',
            # Date formats
            r'(\d{1,2})/(\d{1,2})/(\d{2,4})': lambda m: f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
        }
        
        for pattern, replacement in replacements.items():
            result[f"{text_column}_norm"] = result[f"{text_column}_norm"].str.replace(
                pattern, replacement, regex=True
            )
            
    # Advanced processing for all locales
    # Standardize whitespace
    result[f"{text_column}_norm"] = result[f"{text_column}_norm"].str.replace(r'\s+', ' ', regex=True)
    
    # Standardize case (maintain proper case for proper nouns)
    def smart_case(text):
        if text.isupper():
            return text.title()
        return text
    
    result[f"{text_column}_norm"] = result[f"{text_column}_norm"].apply(smart_case)
    
    return result

def peek_all_sheets(file_path: Union[str, Path], rows: int = 5) -> Dict[str, pd.DataFrame]:
    """
    Peek at the content of all sheets in an Excel file without
    making assumptions about headers.
    
    Args:
        file_path: Path to the Excel file
        rows: Number of rows to preview from each sheet
        
    Returns:
        Dict[str, pd.DataFrame]: Dictionary mapping sheet names to preview DataFrames
    """
    file_path = Path(file_path)
    engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
    
    try:
        excel_file = pd.ExcelFile(file_path, engine=engine)
    except Exception as e:
        # Try alternate engine
        alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
        try:
            excel_file = pd.ExcelFile(file_path, alt_engine)
            engine = alt_engine
        except Exception as e2:
            return {
                "error": f"Failed to open with both engines: {str(e)} / {str(e2)}"
            }
    
    result = {}
    for sheet in excel_file.sheet_names:
        try:
            # Read with no header assumptions
            df = pd.read_excel(file_path, sheet_name=sheet, header=None, nrows=rows, engine=engine)
            result[sheet] = df
        except Exception as e:
            result[f"{sheet}_error"] = str(e)
            
    return result

def analyze_sheet_differences(directory: Union[str, Path], sample_size: int = 5) -> Dict:
    """
    Analyze differences between sheets across multiple Excel files.
    Helps identify how sheet names, structures, and headers vary.
    
    Args:
        directory: Directory containing Excel files
        sample_size: Number of files to sample if there are many
        
    Returns:
        Dict: Analysis of sheet differences
    """
    directory = Path(directory)
    excel_files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    # Sample files if there are too many
    if len(excel_files) > sample_size:
        import random
        excel_files = random.sample(excel_files, sample_size)
    
    result = {
        "file_count": len(excel_files),
        "files_analyzed": [f.name for f in excel_files],
        "sheet_names": {},
        "year_patterns": {},
    }
    
    all_sheet_names = set()
    
    for file_path in excel_files:
        # Extract year from filename if possible
        year_match = None
        for part in file_path.stem.split("_"):
            if part.isdigit() and len(part) == 4:
                year_match = part
                break
                
        info = inspect_excel_structure(file_path)
        sheets = info.get("sheets", [])
        
        # Track all unique sheet names
        all_sheet_names.update(sheets)
        
        # Group by year
        if year_match:
            if year_match not in result["year_patterns"]:
                result["year_patterns"][year_match] = {"files": [], "sheets": set()}
            
            result["year_patterns"][year_match]["files"].append(file_path.name)
            result["year_patterns"][year_match]["sheets"].update(sheets)
        
        # Track sheet name occurrence
        for sheet in sheets:
            if sheet not in result["sheet_names"]:
                result["sheet_names"][sheet] = {"count": 0, "files": []}
            
            result["sheet_names"][sheet]["count"] += 1
            result["sheet_names"][sheet]["files"].append(file_path.name)
    
    # Calculate sheet name statistics
    result["unique_sheet_names"] = list(all_sheet_names)
    result["total_unique_sheets"] = len(all_sheet_names)
    
    # Find potential matches between different sheet names (for mapping)
    potential_matches = {}
    
    # Common variations to look for
    variations = {
        "End of Concession": ["EOC", "Concession", "Concesión", "Concesion"],
        "End of Life": ["EOL", "Life", "Vida Útil", "Vida Util"]
    }
    
    for target, variants in variations.items():
        potential_matches[target] = []
        for sheet in all_sheet_names:
            sheet_lower = sheet.lower()
            if any(var.lower() in sheet_lower for var in variants):
                potential_matches[target].append(sheet)
    
    result["potential_sheet_mappings"] = potential_matches
    
    return result

def analyze_header_patterns(directory: Union[str, Path], 
                            target_sheets: Optional[List[str]] = None,
                            sample_size: int = 5) -> Dict:
    """
    Analyze header patterns across Excel files to understand 
    how the header row position and column names vary.
    
    Args:
        directory: Directory containing Excel files
        target_sheets: Optional list of sheet names to analyze
        sample_size: Number of files to sample if there are many
        
    Returns:
        Dict: Analysis of header patterns
    """
    from ArgGIS.pipeline.utils.transform_utils import detect_header_row
    
    directory = Path(directory)
    excel_files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    # Sample files if there are too many
    if len(excel_files) > sample_size:
        import random
        excel_files = random.sample(excel_files, sample_size)
    
    result = {
        "files_analyzed": [f.name for f in excel_files],
        "header_patterns": {},
        "unique_columns": set(),
        "column_occurrences": {},
        "header_row_positions": {}
    }
    
    for file_path in excel_files:
        try:
            # Determine engine
            engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
            
            # Get sheet names
            try:
                excel_file = pd.ExcelFile(file_path, engine=engine)
                sheets = excel_file.sheet_names
            except Exception:
                alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
                excel_file = pd.ExcelFile(file_path, engine=alt_engine)
                sheets = excel_file.sheet_names
                engine = alt_engine
            
            # Filter to target sheets if specified
            if target_sheets:
                sheets = [s for s in sheets if any(t.lower() in s.lower() for t in target_sheets)]
            
            for sheet in sheets:
                # Read without headers first
                df_raw = pd.read_excel(file_path, sheet_name=sheet, header=None, engine=engine)
                
                # Try to detect header row
                header_row = detect_header_row(df_raw)
                
                # Store header row position
                position_key = f"{header_row}"
                if position_key not in result["header_row_positions"]:
                    result["header_row_positions"][position_key] = {"count": 0, "files": []}
                
                result["header_row_positions"][position_key]["count"] += 1
                result["header_row_positions"][position_key]["files"].append(
                    f"{file_path.name}::{sheet}"
                )
                
                # Now read with detected header
                df = pd.read_excel(file_path, sheet_name=sheet, header=header_row, engine=engine)
                
                # Track column patterns
                columns = list(df.columns)
                column_types = {col: str(df[col].dtype) for col in columns}
                
                # Get first few values for each column
                column_samples = {}
                for col in columns:
                    values = df[col].dropna().head(3).tolist()
                    column_samples[col] = [str(v) for v in values]
                
                # Store pattern
                pattern_key = f"{file_path.name}::{sheet}"
                result["header_patterns"][pattern_key] = {
                    "header_row": header_row,
                    "column_count": len(columns),
                    "columns": columns,
                    "column_types": column_types,
                    "column_samples": column_samples
                }
                
                # Track unique columns
                result["unique_columns"].update(columns)
                
                # Track column occurrences
                for col in columns:
                    if col not in result["column_occurrences"]:
                        result["column_occurrences"][col] = {"count": 0, "files": []}
                    
                    result["column_occurrences"][col]["count"] += 1
                    result["column_occurrences"][col]["files"].append(
                        f"{file_path.name}::{sheet}"
                    )
        
        except Exception as e:
            logger.error(f"Error analyzing {file_path.name}: {e}")
    
    # Convert unique columns to list for JSON serialization
    result["unique_columns"] = list(result["unique_columns"])
    
    return result

def generate_updated_manifest(directory: Union[str, Path], 
                             output_path: Union[str, Path],
                             analysis_result: Optional[Dict] = None) -> pd.DataFrame:
    """
    Generate an updated manifest file with correct sheet mappings based on analysis.
    
    Args:
        directory: Directory containing Excel files
        output_path: Path to save the updated manifest
        analysis_result: Optional pre-computed analysis from analyze_sheet_differences
        
    Returns:
        DataFrame: Updated manifest DataFrame
    """
    if analysis_result is None:
        analysis_result = analyze_sheet_differences(directory)
    
    directory = Path(directory)
    output_path = Path(output_path)
    
    # Extract potential mappings from analysis
    potential_mappings = analysis_result.get("potential_sheet_mappings", {})
    
    # Create mapping dictionaries
    eoc_sheets = set(potential_mappings.get("End of Concession", []))
    eol_sheets = set(potential_mappings.get("End of Life", []))
    
    # Find all Excel files
    excel_files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    # Prepare manifest data
    manifest_data = []
    
    for file_path in excel_files:
        try:
            # Determine engine
            engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
            
            # Extract year from filename
            year = None
            for part in file_path.stem.split("_"):
                if part.isdigit() and len(part) == 4:
                    year = int(part)
                    break
            
            # Get sheet names
            try:
                excel_file = pd.ExcelFile(file_path, engine=engine)
                sheets = excel_file.sheet_names
            except Exception:
                alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
                try:
                    excel_file = pd.ExcelFile(file_path, engine=alt_engine)
                    sheets = excel_file.sheet_names
                    engine = alt_engine
                except Exception as e:
                    logger.error(f"Cannot read {file_path.name}: {e}")
                    continue
            
            # Map sheets to scope
            for sheet in sheets:
                scope = None
                if sheet in eoc_sheets:
                    scope = "end_of_concession"
                elif sheet in eol_sheets:
                    scope = "end_of_life"
                else:
                    # Try direct matching
                    if "concession" in sheet.lower() or "eoc" in sheet.lower().replace(" ", ""):
                        scope = "end_of_concession"
                    elif "life" in sheet.lower() or "eol" in sheet.lower().replace(" ", ""):
                        scope = "end_of_life"
                    else:
                        scope = "unknown"
                
                manifest_data.append({
                    'filename': file_path.name,
                    'year': year,
                    'sheet_name': sheet,
                    'scope': scope,
                    'engine': engine
                })
                
        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {e}")
    
    # Create DataFrame and save
    manifest_df = pd.DataFrame(manifest_data)
    manifest_df.to_csv(output_path, index=False)
    
    logger.info(f"Generated manifest with {len(manifest_df)} entries at {output_path}")
    
    return manifest_df

def create_sheet_mapping_report(directory: Union[str, Path], output_path: Union[str, Path]) -> Dict:
    """
    Create a detailed report on sheet name variations across all Excel files.
    This helps in understanding how to map similar sheets across different file formats.
    
    Args:
        directory: Directory containing Excel files
        output_path: Path to save the report
        
    Returns:
        Dict: Report details
    """
    directory = Path(directory)
    output_path = Path(output_path)
    
    # Get sheet analysis
    analysis = analyze_sheet_differences(directory, sample_size=1000)  # Analyze all files
    
    # Group files by year
    year_groups = {}
    for file_path in directory.glob("*.xls*"):
        year = None
        for part in file_path.stem.split("_"):
            if part.isdigit() and len(part) == 4:
                year = part
                break
        
        if year:
            if year not in year_groups:
                year_groups[year] = []
            year_groups[year].append(file_path.name)
    
    # Build a mapping of how sheet names change over the years
    sheet_evolution = {}
    
    # Identify patterns for EOC and EOL
    eoc_patterns = ["concession", "eoc", "conc"]
    eol_patterns = ["life", "eol", "vida"]
    
    for year, files in sorted(year_groups.items()):
        sheet_evolution[year] = {"files": files, "eoc_sheets": [], "eol_sheets": []}
        
        for file in files:
            try:
                info = inspect_excel_structure(directory / file)
                
                for sheet in info.get("sheets", []):
                    sheet_lower = sheet.lower()
                    
                    if any(pattern in sheet_lower.replace(" ", "") for pattern in eoc_patterns):
                        sheet_evolution[year]["eoc_sheets"].append(sheet)
                    
                    if any(pattern in sheet_lower.replace(" ", "") for pattern in eol_patterns):
                        sheet_evolution[year]["eol_sheets"].append(sheet)
            except:
                pass
    
    # Remove duplicates
    for year_data in sheet_evolution.values():
        year_data["eoc_sheets"] = list(set(year_data["eoc_sheets"]))
        year_data["eol_sheets"] = list(set(year_data["eol_sheets"]))
    
    # Generate report
    report = {
        "sheet_analysis": analysis,
        "year_groups": year_groups,
        "sheet_evolution": sheet_evolution,
        "recommendations": {
            "eoc_mappings": {},
            "eol_mappings": {}
        }
    }
    
    # Generate recommendations
    for year, data in sheet_evolution.items():
        if data["eoc_sheets"]:
            report["recommendations"]["eoc_mappings"][year] = data["eoc_sheets"][0]
        if data["eol_sheets"]:
            report["recommendations"]["eol_mappings"][year] = data["eol_sheets"][0]
    
    # Save report
    with open(output_path, 'w') as f:
        json.dump(report, f, indent=2)
    
    logger.info(f"Sheet mapping report saved to {output_path}")
    
    return report

def preview_file_structure(file_path: Union[str, Path]) -> Dict:
    """
    Create a detailed preview of an Excel file's structure including sheet layout
    and cell formatting information.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dict: Detailed file structure information
    """
    file_path = Path(file_path)
    result = {
        "filename": file_path.name,
        "sheets": {}
    }
    
    # Determine engine
    engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
    
    try:
        if engine == 'openpyxl':
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Collect sheet information
                sheet_info = {
                    "dimensions": f"{max_row} rows x {max_col} columns",
                    "first_rows": [],
                    "merged_cells": [str(cell_range) for cell_range in sheet.merged_cells],
                    "has_header_formatting": False
                }
                
                # Check for header formatting (usually bold or different background)
                if max_row > 0 and max_col > 0:
                    header_format_indicators = 0
                    
                    # Check first few cells in first few rows
                    for r in range(1, min(6, max_row + 1)):
                        row_data = []
                        
                        for c in range(1, min(10, max_col + 1)):
                            cell = sheet.cell(r, c)
                            value = cell.value
                            
                            # Check for formatting in potential header rows
                            if r <= 3:
                                if cell.font and cell.font.bold:
                                    header_format_indicators += 1
                                if cell.fill and cell.fill.patternType != 'none':
                                    header_format_indicators += 1
                            
                            row_data.append(str(value) if value is not None else None)
                        
                        sheet_info["first_rows"].append(row_data)
                    
                    sheet_info["has_header_formatting"] = header_format_indicators > 2
                
                result["sheets"][sheet_name] = sheet_info
                
        else:  # xlrd
            import xlrd
            wb = xlrd.open_workbook(file_path, formatting_info=False)
            
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                
                # Get dimensions
                rows = sheet.nrows
                cols = sheet.ncols
                
                # Collect sheet information
                sheet_info = {
                    "dimensions": f"{rows} rows x {cols} columns",
                    "first_rows": [],
                    "has_header_formatting": False  # xlrd can't easily check formatting
                }
                
                # Get first few rows
                for r in range(min(5, rows)):
                    row_data = []
                    for c in range(min(10, cols)):
                        value = sheet.cell_value(r, c)
                        row_data.append(str(value) if value else None)
                    
                    sheet_info["first_rows"].append(row_data)
                
                result["sheets"][sheet_name] = sheet_info
    
    except Exception as e:
        result["error"] = str(e)
    
    return result

def interactive_excel_explorer(file_path: Union[str, Path], save_report_path: Optional[str] = None) -> Dict:
    """
    Comprehensive interactive Excel explorer that provides detailed information
    about the structure and content of an Excel file.
    
    Args:
        file_path: Path to the Excel file
        save_report_path: Optional path to save the report
        
    Returns:
        Dict: Comprehensive analysis of the Excel file
    """
    from ArgGIS.pipeline.utils.transform_utils import detect_header_row
    
    file_path = Path(file_path)
    engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
    
    # Basic file info
    result = {
        "filename": file_path.name,
        "file_size_kb": file_path.stat().st_size / 1024,
        "timestamp": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
        "sheets": {}
    }
    
    try:
        # Try to open with primary engine
        try:
            excel_file = pd.ExcelFile(file_path, engine=engine)
        except Exception:
            # Try alternate engine
            alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
            excel_file = pd.ExcelFile(file_path, engine=alt_engine)
            engine = alt_engine
            
        result["engine_used"] = engine
        
        # Analyze each sheet
        for sheet_name in excel_file.sheet_names:
            # Try to detect header row
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine=engine)
            header_row = detect_header_row(df_raw)
            
            # Read with detected header
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, engine=engine)
            
            # Get column info
            columns = list(df.columns)
            col_types = {str(col): str(df[col].dtype) for col in columns}
            
            # Check for empty columns
            empty_cols = [col for col in columns if df[col].isna().all()]
            
            # Check for potential ID columns
            potential_id_cols = []
            for col in columns:
                if df[col].dtype in ['object', 'string']:
                    # String columns with low cardinality might be categories/IDs
                    unique_ratio = df[col].nunique() / len(df) if len(df) > 0 else 0
                    if 0 < unique_ratio <= 0.5:  # Heuristic for ID columns
                        potential_id_cols.append(str(col))
            
            # Detect potential value columns
            numeric_cols = [col for col in columns if pd.api.types.is_numeric_dtype(df[col].dtype)]
            
            # Calculate basic stats
            stats = {
                "row_count": len(df),
                "column_count": len(columns),
                "header_row": header_row,
                "empty_cells_count": df.isna().sum().sum(),
                "empty_rows_count": sum(df.isna().all(axis=1)),
                "empty_columns": empty_cols,
                "potential_id_columns": potential_id_cols,
                "numeric_columns": numeric_cols,
                "column_types": col_types,
                "first_row_after_header": df.iloc[0].to_dict() if not df.empty else {},
                "column_samples": {}
            }
            
            # Get samples for each column
            for col in columns:
                non_null_values = df[col].dropna()
                if len(non_null_values) > 0:
                    stats["column_samples"][str(col)] = [
                        str(val) for val in non_null_values.head(3).tolist()
                    ]
            
            result["sheets"][sheet_name] = stats
        
        # Additional metadata for common sheet patterns
        sheet_patterns = {
            "eoc_sheets": [],
            "eol_sheets": []
        }
        
        for sheet in excel_file.sheet_names:
            sheet_lower = sheet.lower().replace(" ", "")
            if any(pattern in sheet_lower for pattern in ["concession", "eoc", "conc"]):
                sheet_patterns["eoc_sheets"].append(sheet)
            if any(pattern in sheet_lower for pattern in ["life", "eol", "vida"]):
                sheet_patterns["eol_sheets"].append(sheet)
        
        result["sheet_patterns"] = sheet_patterns
        
        # Save report if requested
        if save_report_path:
            with open(save_report_path, 'w') as f:
                json.dump(result, f, indent=2)
            print(f"Report saved to {save_report_path}")
        
        return result
    
    except Exception as e:
        result["error"] = str(e)
        return result
    
def explore_excel_files(directory: str) -> Dict:
    """
    Directly explore Excel files and list all their sheet names.
    No fancy algorithms, just straight examination of what's in each file.
    
    Args:
        directory: Directory containing the Excel files
    
    Returns:
        Dict: Dictionary with file details and sheet names
    """
    directory = Path(directory)
    files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    results = {
        "total_files": len(files),
        "file_details": []
    }
    
    for file_path in files:
        try:
            # Try both engines for maximum compatibility
            engines = ['openpyxl', 'xlrd']
            sheets = []
            used_engine = None
            
            for engine in engines:
                try:
                    excel = pd.ExcelFile(file_path, engine=engine)
                    sheets = excel.sheet_names
                    used_engine = engine
                    break
                except Exception as e:
                    logger.warning(f"Failed with engine {engine} for {file_path.name}: {e}")
            
            if not sheets:
                logger.error(f"Could not read {file_path.name} with any engine")
                continue
                
            # Extract year from filename
            year = None
            for part in file_path.stem.split('_'):
                if part.isdigit() and len(part) == 4:
                    year = part
                    break
            
            file_info = {
                "filename": file_path.name,
                "year": year,
                "sheets": sheets,
                "engine_used": used_engine
            }
            
            results["file_details"].append(file_info)
            logger.info(f"Processed {file_path.name}: found {len(sheets)} sheets")
            
        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {e}")
    
    return results

def print_sheet_names_by_year(exploration_results: Dict) -> None:
    """
    Print sheet names grouped by year for easy pattern recognition.
    
    Args:
        exploration_results: Results from explore_excel_files function
    """
    # Group by year
    years = {}
    for file_info in exploration_results["file_details"]:
        year = file_info.get("year", "unknown")
        if year not in years:
            years[year] = []
        
        # Add sheet names with file info
        for sheet in file_info["sheets"]:
            years[year].append({
                "filename": file_info["filename"],
                "sheet": sheet
            })
    
    # Print by year
    print("\n=== SHEET NAMES BY YEAR ===")
    for year in sorted(years.keys()):
        print(f"\nYear: {year}")
        for item in years[year]:
            print(f"  {item['filename']} -> {item['sheet']}")

def build_sheet_mapping(exploration_results: Dict) -> Dict:
    """
    Build a mapping between sheet names and their likely type (EOC or EOL).
    
    Args:
        exploration_results: Results from explore_excel_files function
    
    Returns:
        Dict: Mapping of sheet names to types
    """
    sheet_mapping = {}
    
    # Keywords to identify sheet types
    eoc_keywords = ["concession", "eoc", "conc", "concesión", "concesion"]
    eol_keywords = ["life", "eol", "vida", "útil", "util"]
    
    # Check each file's sheets
    for file_info in exploration_results["file_details"]:
        for sheet in file_info["sheets"]:
            sheet_lower = sheet.lower()
            
            # Check if sheet matches EOC pattern
            if any(keyword in sheet_lower.replace(" ", "") for keyword in eoc_keywords):
                sheet_mapping[sheet] = "end_of_concession"
            
            # Check if sheet matches EOL pattern
            elif any(keyword in sheet_lower.replace(" ", "") for keyword in eol_keywords):
                sheet_mapping[sheet] = "end_of_life"
            
            # Unknown type
            else:
                sheet_mapping[sheet] = "unknown"
    
    return sheet_mapping

def create_fixed_manifest(directory: str, output_path: str) -> None:
    """
    Create a fixed manifest.csv file based on the actual sheets in the Excel files.
    
    Args:
        directory: Directory containing the Excel files
        output_path: Path to save the fixed manifest.csv
    """
    directory = Path(directory)
    
    # Explore Excel files
    results = explore_excel_files(directory)
    
    # Build sheet type mapping
    sheet_mapping = build_sheet_mapping(results)
    
    # Create manifest data
    manifest_data = []
    
    for file_info in results["file_details"]:
        filename = file_info["filename"]
        year = file_info.get("year")
        engine = file_info.get("engine_used", "openpyxl")
        
        for sheet in file_info["sheets"]:
            scope = sheet_mapping.get(sheet, "unknown")
            
            manifest_data.append({
                "filename": filename,
                "year": year,
                "sheet_name": sheet,
                "scope": scope,
                "engine": engine
            })
    
    # Create DataFrame and save to CSV
    manifest_df = pd.DataFrame(manifest_data)
    manifest_df.to_csv(output_path, index=False)
    
    logger.info(f"Created fixed manifest with {len(manifest_df)} entries at {output_path}")
    print(f"Fixed manifest saved to {output_path}")

def preview_excel_content(file_path: str, num_rows: int = 5) -> Dict:
    """
    Preview the content of all sheets in an Excel file.
    
    Args:
        file_path: Path to the Excel file
        num_rows: Number of rows to preview from each sheet
    
    Returns:
        Dict: Dictionary with preview data for each sheet
    """
    file_path = Path(file_path)
    
    # Try both engines
    for engine in ['openpyxl', 'xlrd']:
        try:
            excel = pd.ExcelFile(file_path, engine=engine)
            break
        except Exception:
            continue
    else:
        return {"error": f"Could not open {file_path.name} with any engine"}
    
    preview_data = {}
    
    for sheet in excel.sheet_names:
        try:
            # Try first with no header
            df_no_header = pd.read_excel(file_path, sheet_name=sheet, header=None, nrows=num_rows, engine=engine)
            preview_data[f"{sheet}_raw"] = df_no_header.to_dict(orient='records')
            
            # Try to identify header row (simple approach - look at row 2, 3, 4)
            for header_row in [0, 1, 2, 3, 4]:
                if header_row < len(df_no_header):
                    df_with_header = pd.read_excel(
                        file_path, sheet_name=sheet, 
                        header=header_row, nrows=num_rows, engine=engine
                    )
                    preview_data[f"{sheet}_header{header_row}"] = df_with_header.to_dict(orient='records')
        except Exception as e:
            preview_data[f"{sheet}_error"] = str(e)
    
    return preview_data


import os
import pandas as pd
import json
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Union
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def explore_excel_files(directory: str) -> Dict:
    """
    Directly explore Excel files and list all their sheet names.
    No fancy algorithms, just straight examination of what's in each file.
    
    Args:
        directory: Directory containing the Excel files
    
    Returns:
        Dict: Dictionary with file details and sheet names
    """
    directory = Path(directory)
    files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    results = {
        "total_files": len(files),
        "file_details": []
    }
    
    for file_path in files:
        try:
            # Try both engines for maximum compatibility
            engines = ['openpyxl', 'xlrd']
            sheets = []
            used_engine = None
            
            for engine in engines:
                try:
                    excel = pd.ExcelFile(file_path, engine=engine)
                    sheets = excel.sheet_names
                    used_engine = engine
                    break
                except Exception as e:
                    logger.warning(f"Failed with engine {engine} for {file_path.name}: {e}")
            
            if not sheets:
                logger.error(f"Could not read {file_path.name} with any engine")
                continue
                
            # Extract year from filename
            year = None
            for part in file_path.stem.split('_'):
                if part.isdigit() and len(part) == 4:
                    year = part
                    break
            
            file_info = {
                "filename": file_path.name,
                "year": year,
                "sheets": sheets,
                "engine_used": used_engine
            }
            
            results["file_details"].append(file_info)
            logger.info(f"Processed {file_path.name}: found {len(sheets)} sheets")
            
        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {e}")
    
    return results

def print_sheet_names_by_year(exploration_results: Dict) -> None:
    """
    Print sheet names grouped by year for easy pattern recognition.
    
    Args:
        exploration_results: Results from explore_excel_files function
    """
    # Group by year
    years = {}
    for file_info in exploration_results["file_details"]:
        year = file_info.get("year", "unknown")
        if year not in years:
            years[year] = []
        
        # Add sheet names with file info
        for sheet in file_info["sheets"]:
            years[year].append({
                "filename": file_info["filename"],
                "sheet": sheet
            })
    
    # Print by year
    print("\n=== SHEET NAMES BY YEAR ===")
    for year in sorted(years.keys()):
        print(f"\nYear: {year}")
        for item in years[year]:
            print(f"  {item['filename']} -> {item['sheet']}")

def build_sheet_mapping(exploration_results: Dict) -> Dict:
    """
    Build a mapping between sheet names and their likely type (EOC or EOL).
    
    Args:
        exploration_results: Results from explore_excel_files function
    
    Returns:
        Dict: Mapping of sheet names to types
    """
    sheet_mapping = {}
    
    # Keywords to identify sheet types
    eoc_keywords = ["concession", "eoc", "conc", "concesión", "concesion"]
    eol_keywords = ["life", "eol", "vida", "útil", "util"]
    
    # Check each file's sheets
    for file_info in exploration_results["file_details"]:
        for sheet in file_info["sheets"]:
            sheet_lower = sheet.lower()
            
            # Check if sheet matches EOC pattern
            if any(keyword in sheet_lower.replace(" ", "") for keyword in eoc_keywords):
                sheet_mapping[sheet] = "end_of_concession"
            
            # Check if sheet matches EOL pattern
            elif any(keyword in sheet_lower.replace(" ", "") for keyword in eol_keywords):
                sheet_mapping[sheet] = "end_of_life"
            
            # Unknown type
            else:
                sheet_mapping[sheet] = "unknown"
    
    return sheet_mapping

def create_fixed_manifest(directory: str, output_path: str) -> None:
    """
    Create a fixed manifest.csv file based on the actual sheets in the Excel files.
    
    Args:
        directory: Directory containing the Excel files
        output_path: Path to save the fixed manifest.csv
    """
    directory = Path(directory)
    
    # Explore Excel files
    results = explore_excel_files(directory)
    
    # Build sheet type mapping
    sheet_mapping = build_sheet_mapping(results)
    
    # Create manifest data
    manifest_data = []
    
    for file_info in results["file_details"]:
        filename = file_info["filename"]
        year = file_info.get("year")
        engine = file_info.get("engine_used", "openpyxl")
        
        for sheet in file_info["sheets"]:
            scope = sheet_mapping.get(sheet, "unknown")
            
            manifest_data.append({
                "filename": filename,
                "year": year,
                "sheet_name": sheet,
                "scope": scope,
                "engine": engine
            })
    
    # Create DataFrame and save to CSV
    manifest_df = pd.DataFrame(manifest_data)
    manifest_df.to_csv(output_path, index=False)
    
    logger.info(f"Created fixed manifest with {len(manifest_df)} entries at {output_path}")
    print(f"Fixed manifest saved to {output_path}")

def preview_excel_content(file_path: str, num_rows: int = 5) -> Dict:
    """
    Preview the content of all sheets in an Excel file.
    
    Args:
        file_path: Path to the Excel file
        num_rows: Number of rows to preview from each sheet
    
    Returns:
        Dict: Dictionary with preview data for each sheet
    """
    file_path = Path(file_path)
    
    # Try both engines
    for engine in ['openpyxl', 'xlrd']:
        try:
            excel = pd.ExcelFile(file_path, engine=engine)
            break
        except Exception:
            continue
    else:
        return {"error": f"Could not open {file_path.name} with any engine"}
    
    preview_data = {}
    
    for sheet in excel.sheet_names:
        try:
            # Try first with no header
            df_no_header = pd.read_excel(file_path, sheet_name=sheet, header=None, nrows=num_rows, engine=engine)
            preview_data[f"{sheet}_raw"] = df_no_header.to_dict(orient='records')
            
            # Try to identify header row (simple approach - look at row 2, 3, 4)
            for header_row in [0, 1, 2, 3, 4]:
                if header_row < len(df_no_header):
                    df_with_header = pd.read_excel(
                        file_path, sheet_name=sheet, 
                        header=header_row, nrows=num_rows, engine=engine
                    )
                    preview_data[f"{sheet}_header{header_row}"] = df_with_header.to_dict(orient='records')
        except Exception as e:
            preview_data[f"{sheet}_error"] = str(e)
    
    return preview_data
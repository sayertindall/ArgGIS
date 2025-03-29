import os
import logging
import pandas as pd
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union
from datetime import datetime

# Assuming your existing utilities are imported like this:
# from ArgGIS.pipeline.utils.transform_utils import detect_header_row, normalize_columns, tidy_reserve_table

logger = logging.getLogger(__name__)

def inspect_excel_structure(file_path: Union[str, Path]) -> Dict:
    """
    Inspect an Excel file's structure without loading full content.
    Shows available sheets and basic stats.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dict: Information about the Excel file structure
    """
    file_path = Path(file_path)
    result = {
        "filename": file_path.name,
        "path": str(file_path.absolute()),
        "file_size_kb": file_path.stat().st_size / 1024,
        "sheets": [],
    }
    
    # Try with different engines
    for engine in ['openpyxl', 'xlrd']:
        try:
            excel_file = pd.ExcelFile(file_path, engine=engine)
            result["engine_used"] = engine
            result["sheets"] = excel_file.sheet_names
            break
        except Exception as e:
            result["error"] = f"Error with engine {engine}: {str(e)}"
    
    return result

def peek_all_sheets(file_path: Union[str, Path], rows: int = 5) -> Dict[str, pd.DataFrame]:
    """
    Peek at the content of all sheets in an Excel file without
    making assumptions about headers.
    
    Args:
        file_path: Path to the Excel file
        rows: Number of rows to preview from each sheet
        
    Returns:
        Dict[str, pd.DataFrame]: Dictionary mapping sheet names to preview DataFrames
    """
    file_path = Path(file_path)
    engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
    
    try:
        excel_file = pd.ExcelFile(file_path, engine=engine)
    except Exception as e:
        # Try alternate engine
        alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
        try:
            excel_file = pd.ExcelFile(file_path, alt_engine)
            engine = alt_engine
        except Exception as e2:
            return {
                "error": f"Failed to open with both engines: {str(e)} / {str(e2)}"
            }
    
    result = {}
    for sheet in excel_file.sheet_names:
        try:
            # Read with no header assumptions
            df = pd.read_excel(file_path, sheet_name=sheet, header=None, nrows=rows, engine=engine)
            result[sheet] = df
        except Exception as e:
            result[f"{sheet}_error"] = str(e)
            
    return result

def analyze_sheet_differences(directory: Union[str, Path], sample_size: int = 5) -> Dict:
    """
    Analyze differences between sheets across multiple Excel files.
    Helps identify how sheet names, structures, and headers vary.
    
    Args:
        directory: Directory containing Excel files
        sample_size: Number of files to sample if there are many
        
    Returns:
        Dict: Analysis of sheet differences
    """
    directory = Path(directory)
    excel_files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    # Sample files if there are too many
    if len(excel_files) > sample_size:
        import random
        excel_files = random.sample(excel_files, sample_size)
    
    result = {
        "file_count": len(excel_files),
        "files_analyzed": [f.name for f in excel_files],
        "sheet_names": {},
        "year_patterns": {},
    }
    
    all_sheet_names = set()
    
    for file_path in excel_files:
        # Extract year from filename if possible
        year_match = None
        for part in file_path.stem.split("_"):
            if part.isdigit() and len(part) == 4:
                year_match = part
                break
                
        info = inspect_excel_structure(file_path)
        sheets = info.get("sheets", [])
        
        # Track all unique sheet names
        all_sheet_names.update(sheets)
        
        # Group by year
        if year_match:
            if year_match not in result["year_patterns"]:
                result["year_patterns"][year_match] = {"files": [], "sheets": set()}
            
            result["year_patterns"][year_match]["files"].append(file_path.name)
            result["year_patterns"][year_match]["sheets"].update(sheets)
        
        # Track sheet name occurrence
        for sheet in sheets:
            if sheet not in result["sheet_names"]:
                result["sheet_names"][sheet] = {"count": 0, "files": []}
            
            result["sheet_names"][sheet]["count"] += 1
            result["sheet_names"][sheet]["files"].append(file_path.name)
    
    # Calculate sheet name statistics
    result["unique_sheet_names"] = list(all_sheet_names)
    result["total_unique_sheets"] = len(all_sheet_names)
    
    # Find potential matches between different sheet names (for mapping)
    potential_matches = {}
    
    # Common variations to look for
    variations = {
        "End of Concession": ["EOC", "Concession", "Concesión", "Concesion"],
        "End of Life": ["EOL", "Life", "Vida Útil", "Vida Util"]
    }
    
    for target, variants in variations.items():
        potential_matches[target] = []
        for sheet in all_sheet_names:
            sheet_lower = sheet.lower()
            if any(var.lower() in sheet_lower for var in variants):
                potential_matches[target].append(sheet)
    
    result["potential_sheet_mappings"] = potential_matches
    
    return result

def analyze_header_patterns(directory: Union[str, Path], 
                            target_sheets: Optional[List[str]] = None,
                            sample_size: int = 5) -> Dict:
    """
    Analyze header patterns across Excel files to understand 
    how the header row position and column names vary.
    
    Args:
        directory: Directory containing Excel files
        target_sheets: Optional list of sheet names to analyze
        sample_size: Number of files to sample if there are many
        
    Returns:
        Dict: Analysis of header patterns
    """
    from ArgGIS.pipeline.utils.transform_utils import detect_header_row
    
    directory = Path(directory)
    excel_files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    # Sample files if there are too many
    if len(excel_files) > sample_size:
        import random
        excel_files = random.sample(excel_files, sample_size)
    
    result = {
        "files_analyzed": [f.name for f in excel_files],
        "header_patterns": {},
        "unique_columns": set(),
        "column_occurrences": {},
        "header_row_positions": {}
    }
    
    for file_path in excel_files:
        try:
            # Determine engine
            engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
            
            # Get sheet names
            try:
                excel_file = pd.ExcelFile(file_path, engine=engine)
                sheets = excel_file.sheet_names
            except Exception:
                alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
                excel_file = pd.ExcelFile(file_path, engine=alt_engine)
                sheets = excel_file.sheet_names
                engine = alt_engine
            
            # Filter to target sheets if specified
            if target_sheets:
                sheets = [s for s in sheets if any(t.lower() in s.lower() for t in target_sheets)]
            
            for sheet in sheets:
                # Read without headers first
                df_raw = pd.read_excel(file_path, sheet_name=sheet, header=None, engine=engine)
                
                # Try to detect header row
                header_row = detect_header_row(df_raw)
                
                # Store header row position
                position_key = f"{header_row}"
                if position_key not in result["header_row_positions"]:
                    result["header_row_positions"][position_key] = {"count": 0, "files": []}
                
                result["header_row_positions"][position_key]["count"] += 1
                result["header_row_positions"][position_key]["files"].append(
                    f"{file_path.name}::{sheet}"
                )
                
                # Now read with detected header
                df = pd.read_excel(file_path, sheet_name=sheet, header=header_row, engine=engine)
                
                # Track column patterns
                columns = list(df.columns)
                column_types = {col: str(df[col].dtype) for col in columns}
                
                # Get first few values for each column
                column_samples = {}
                for col in columns:
                    values = df[col].dropna().head(3).tolist()
                    column_samples[col] = [str(v) for v in values]
                
                # Store pattern
                pattern_key = f"{file_path.name}::{sheet}"
                result["header_patterns"][pattern_key] = {
                    "header_row": header_row,
                    "column_count": len(columns),
                    "columns": columns,
                    "column_types": column_types,
                    "column_samples": column_samples
                }
                
                # Track unique columns
                result["unique_columns"].update(columns)
                
                # Track column occurrences
                for col in columns:
                    if col not in result["column_occurrences"]:
                        result["column_occurrences"][col] = {"count": 0, "files": []}
                    
                    result["column_occurrences"][col]["count"] += 1
                    result["column_occurrences"][col]["files"].append(
                        f"{file_path.name}::{sheet}"
                    )
        
        except Exception as e:
            logger.error(f"Error analyzing {file_path.name}: {e}")
    
    # Convert unique columns to list for JSON serialization
    result["unique_columns"] = list(result["unique_columns"])
    
    return result

def generate_updated_manifest(directory: Union[str, Path], 
                             output_path: Union[str, Path],
                             analysis_result: Optional[Dict] = None) -> pd.DataFrame:
    """
    Generate an updated manifest file with correct sheet mappings based on analysis.
    
    Args:
        directory: Directory containing Excel files
        output_path: Path to save the updated manifest
        analysis_result: Optional pre-computed analysis from analyze_sheet_differences
        
    Returns:
        DataFrame: Updated manifest DataFrame
    """
    if analysis_result is None:
        analysis_result = analyze_sheet_differences(directory)
    
    directory = Path(directory)
    output_path = Path(output_path)
    
    # Extract potential mappings from analysis
    potential_mappings = analysis_result.get("potential_sheet_mappings", {})
    
    # Create mapping dictionaries
    eoc_sheets = set(potential_mappings.get("End of Concession", []))
    eol_sheets = set(potential_mappings.get("End of Life", []))
    
    # Find all Excel files
    excel_files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    # Prepare manifest data
    manifest_data = []
    
    for file_path in excel_files:
        try:
            # Determine engine
            engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
            
            # Extract year from filename
            year = None
            for part in file_path.stem.split("_"):
                if part.isdigit() and len(part) == 4:
                    year = int(part)
                    break
            
            # Get sheet names
            try:
                excel_file = pd.ExcelFile(file_path, engine=engine)
                sheets = excel_file.sheet_names
            except Exception:
                alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
                try:
                    excel_file = pd.ExcelFile(file_path, engine=alt_engine)
                    sheets = excel_file.sheet_names
                    engine = alt_engine
                except Exception as e:
                    logger.error(f"Cannot read {file_path.name}: {e}")
                    continue
            
            # Map sheets to scope
            for sheet in sheets:
                scope = None
                if sheet in eoc_sheets:
                    scope = "end_of_concession"
                elif sheet in eol_sheets:
                    scope = "end_of_life"
                else:
                    # Try direct matching
                    if "concession" in sheet.lower() or "eoc" in sheet.lower().replace(" ", ""):
                        scope = "end_of_concession"
                    elif "life" in sheet.lower() or "eol" in sheet.lower().replace(" ", ""):
                        scope = "end_of_life"
                    else:
                        scope = "unknown"
                
                manifest_data.append({
                    'filename': file_path.name,
                    'year': year,
                    'sheet_name': sheet,
                    'scope': scope,
                    'engine': engine
                })
                
        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {e}")
    
    # Create DataFrame and save
    manifest_df = pd.DataFrame(manifest_data)
    manifest_df.to_csv(output_path, index=False)
    
    logger.info(f"Generated manifest with {len(manifest_df)} entries at {output_path}")
    
    return manifest_df

def create_sheet_mapping_report(directory: Union[str, Path], output_path: Union[str, Path]) -> Dict:
    """
    Create a detailed report on sheet name variations across all Excel files.
    This helps in understanding how to map similar sheets across different file formats.
    
    Args:
        directory: Directory containing Excel files
        output_path: Path to save the report
        
    Returns:
        Dict: Report details
    """
    directory = Path(directory)
    output_path = Path(output_path)
    
    # Get sheet analysis
    analysis = analyze_sheet_differences(directory, sample_size=1000)  # Analyze all files
    
    # Group files by year
    year_groups = {}
    for file_path in directory.glob("*.xls*"):
        year = None
        for part in file_path.stem.split("_"):
            if part.isdigit() and len(part) == 4:
                year = part
                break
        
        if year:
            if year not in year_groups:
                year_groups[year] = []
            year_groups[year].append(file_path.name)
    
    # Build a mapping of how sheet names change over the years
    sheet_evolution = {}
    
    # Identify patterns for EOC and EOL
    eoc_patterns = ["concession", "eoc", "conc"]
    eol_patterns = ["life", "eol", "vida"]
    
    for year, files in sorted(year_groups.items()):
        sheet_evolution[year] = {"files": files, "eoc_sheets": [], "eol_sheets": []}
        
        for file in files:
            try:
                info = inspect_excel_structure(directory / file)
                
                for sheet in info.get("sheets", []):
                    sheet_lower = sheet.lower()
                    
                    if any(pattern in sheet_lower.replace(" ", "") for pattern in eoc_patterns):
                        sheet_evolution[year]["eoc_sheets"].append(sheet)
                    
                    if any(pattern in sheet_lower.replace(" ", "") for pattern in eol_patterns):
                        sheet_evolution[year]["eol_sheets"].append(sheet)
            except:
                pass
    
    # Remove duplicates
    for year_data in sheet_evolution.values():
        year_data["eoc_sheets"] = list(set(year_data["eoc_sheets"]))
        year_data["eol_sheets"] = list(set(year_data["eol_sheets"]))
    
    # Generate report
    report = {
        "sheet_analysis": analysis,
        "year_groups": year_groups,
        "sheet_evolution": sheet_evolution,
        "recommendations": {
            "eoc_mappings": {},
            "eol_mappings": {}
        }
    }
    
    # Generate recommendations
    for year, data in sheet_evolution.items():
        if data["eoc_sheets"]:
            report["recommendations"]["eoc_mappings"][year] = data["eoc_sheets"][0]
        if data["eol_sheets"]:
            report["recommendations"]["eol_mappings"][year] = data["eol_sheets"][0]
    
    # Save report
    with open(output_path, 'w') as f:
        json.dump(report, f, indent=2)
    
    logger.info(f"Sheet mapping report saved to {output_path}")
    
    return report

def preview_file_structure(file_path: Union[str, Path]) -> Dict:
    """
    Create a detailed preview of an Excel file's structure including sheet layout
    and cell formatting information.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dict: Detailed file structure information
    """
    file_path = Path(file_path)
    result = {
        "filename": file_path.name,
        "sheets": {}
    }
    
    # Determine engine
    engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
    
    try:
        if engine == 'openpyxl':
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Collect sheet information
                sheet_info = {
                    "dimensions": f"{max_row} rows x {max_col} columns",
                    "first_rows": [],
                    "merged_cells": [str(cell_range) for cell_range in sheet.merged_cells],
                    "has_header_formatting": False
                }
                
                # Check for header formatting (usually bold or different background)
                if max_row > 0 and max_col > 0:
                    header_format_indicators = 0
                    
                    # Check first few cells in first few rows
                    for r in range(1, min(6, max_row + 1)):
                        row_data = []
                        
                        for c in range(1, min(10, max_col + 1)):
                            cell = sheet.cell(r, c)
                            value = cell.value
                            
                            # Check for formatting in potential header rows
                            if r <= 3:
                                if cell.font and cell.font.bold:
                                    header_format_indicators += 1
                                if cell.fill and cell.fill.patternType != 'none':
                                    header_format_indicators += 1
                            
                            row_data.append(str(value) if value is not None else None)
                        
                        sheet_info["first_rows"].append(row_data)
                    
                    sheet_info["has_header_formatting"] = header_format_indicators > 2
                
                result["sheets"][sheet_name] = sheet_info
                
        else:  # xlrd
            import xlrd
            wb = xlrd.open_workbook(file_path, formatting_info=False)
            
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                
                # Get dimensions
                rows = sheet.nrows
                cols = sheet.ncols
                
                # Collect sheet information
                sheet_info = {
                    "dimensions": f"{rows} rows x {cols} columns",
                    "first_rows": [],
                    "has_header_formatting": False  # xlrd can't easily check formatting
                }
                
                # Get first few rows
                for r in range(min(5, rows)):
                    row_data = []
                    for c in range(min(10, cols)):
                        value = sheet.cell_value(r, c)
                        row_data.append(str(value) if value else None)
                    
                    sheet_info["first_rows"].append(row_data)
                
                result["sheets"][sheet_name] = sheet_info
    
    except Exception as e:
        result["error"] = str(e)
    
    return result

def interactive_excel_explorer(file_path: Union[str, Path], save_report_path: Optional[str] = None) -> Dict:
    """
    Comprehensive interactive Excel explorer that provides detailed information
    about the structure and content of an Excel file.
    
    Args:
        file_path: Path to the Excel file
        save_report_path: Optional path to save the report
        
    Returns:
        Dict: Comprehensive analysis of the Excel file
    """
    from ArgGIS.pipeline.utils.transform_utils import detect_header_row
    
    file_path = Path(file_path)
    engine = 'openpyxl' if file_path.suffix.lower() == '.xlsx' else 'xlrd'
    
    # Basic file info
    result = {
        "filename": file_path.name,
        "file_size_kb": file_path.stat().st_size / 1024,
        "timestamp": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
        "sheets": {}
    }
    
    try:
        # Try to open with primary engine
        try:
            excel_file = pd.ExcelFile(file_path, engine=engine)
        except Exception:
            # Try alternate engine
            alt_engine = 'xlrd' if engine == 'openpyxl' else 'openpyxl'
            excel_file = pd.ExcelFile(file_path, engine=alt_engine)
            engine = alt_engine
            
        result["engine_used"] = engine
        
        # Analyze each sheet
        for sheet_name in excel_file.sheet_names:
            # Try to detect header row
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine=engine)
            header_row = detect_header_row(df_raw)
            
            # Read with detected header
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, engine=engine)
            
            # Get column info
            columns = list(df.columns)
            col_types = {str(col): str(df[col].dtype) for col in columns}
            
            # Check for empty columns
            empty_cols = [col for col in columns if df[col].isna().all()]
            
            # Check for potential ID columns
            potential_id_cols = []
            for col in columns:
                if df[col].dtype in ['object', 'string']:
                    # String columns with low cardinality might be categories/IDs
                    unique_ratio = df[col].nunique() / len(df) if len(df) > 0 else 0
                    if 0 < unique_ratio <= 0.5:  # Heuristic for ID columns
                        potential_id_cols.append(str(col))
            
            # Detect potential value columns
            numeric_cols = [col for col in columns if pd.api.types.is_numeric_dtype(df[col].dtype)]
            
            # Calculate basic stats
            stats = {
                "row_count": len(df),
                "column_count": len(columns),
                "header_row": header_row,
                "empty_cells_count": df.isna().sum().sum(),
                "empty_rows_count": sum(df.isna().all(axis=1)),
                "empty_columns": empty_cols,
                "potential_id_columns": potential_id_cols,
                "numeric_columns": numeric_cols,
                "column_types": col_types,
                "first_row_after_header": df.iloc[0].to_dict() if not df.empty else {},
                "column_samples": {}
            }
            
            # Get samples for each column
            for col in columns:
                non_null_values = df[col].dropna()
                if len(non_null_values) > 0:
                    stats["column_samples"][str(col)] = [
                        str(val) for val in non_null_values.head(3).tolist()
                    ]
            
            result["sheets"][sheet_name] = stats
        
        # Additional metadata for common sheet patterns
        sheet_patterns = {
            "eoc_sheets": [],
            "eol_sheets": []
        }
        
        for sheet in excel_file.sheet_names:
            sheet_lower = sheet.lower().replace(" ", "")
            if any(pattern in sheet_lower for pattern in ["concession", "eoc", "conc"]):
                sheet_patterns["eoc_sheets"].append(sheet)
            if any(pattern in sheet_lower for pattern in ["life", "eol", "vida"]):
                sheet_patterns["eol_sheets"].append(sheet)
        
        result["sheet_patterns"] = sheet_patterns
        
        # Save report if requested
        if save_report_path:
            with open(save_report_path, 'w') as f:
                json.dump(result, f, indent=2)
            print(f"Report saved to {save_report_path}")
        
        return result
    
    except Exception as e:
        result["error"] = str(e)
        return result
    
def explore_excel_files(directory: str) -> Dict:
    """
    Directly explore Excel files and list all their sheet names.
    No fancy algorithms, just straight examination of what's in each file.
    
    Args:
        directory: Directory containing the Excel files
    
    Returns:
        Dict: Dictionary with file details and sheet names
    """
    directory = Path(directory)
    files = list(directory.glob("*.xls")) + list(directory.glob("*.xlsx"))
    
    results = {
        "total_files": len(files),
        "file_details": []
    }
    
    for file_path in files:
        try:
            # Try both engines for maximum compatibility
            engines = ['openpyxl', 'xlrd']
            sheets = []
            used_engine = None
            
            for engine in engines:
                try:
                    excel = pd.ExcelFile(file_path, engine=engine)
                    sheets = excel.sheet_names
                    used_engine = engine
                    break
                except Exception as e:
                    logger.warning(f"Failed with engine {engine} for {file_path.name}: {e}")
            
            if not sheets:
                logger.error(f"Could not read {file_path.name} with any engine")
                continue
                
            # Extract year from filename
            year = None
            for part in file_path.stem.split('_'):
                if part.isdigit() and len(part) == 4:
                    year = part
                    break
            
            file_info = {
                "filename": file_path.name,
                "year": year,
                "sheets": sheets,
                "engine_used": used_engine
            }
            
            results["file_details"].append(file_info)
            logger.info(f"Processed {file_path.name}: found {len(sheets)} sheets")
            
        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {e}")
    
    return results

def print_sheet_names_by_year(exploration_results: Dict) -> None:
    """
    Print sheet names grouped by year for easy pattern recognition.
    
    Args:
        exploration_results: Results from explore_excel_files function
    """
    # Group by year
    years = {}
    for file_info in exploration_results["file_details"]:
        year = file_info.get("year", "unknown")
        if year not in years:
            years[year] = []
        
        # Add sheet names with file info
        for sheet in file_info["sheets"]:
            years[year].append({
                "filename": file_info["filename"],
                "sheet": sheet
            })
    
    # Print by year
    print("\n=== SHEET NAMES BY YEAR ===")
    for year in sorted(years.keys()):
        print(f"\nYear: {year}")
        for item in years[year]:
            print(f"  {item['filename']} -> {item['sheet']}")

def build_sheet_mapping(exploration_results: Dict) -> Dict:
    """
    Build a mapping between sheet names and their likely type (EOC or EOL).
    
    Args:
        exploration_results: Results from explore_excel_files function
    
    Returns:
        Dict: Mapping of sheet names to types
    """
    sheet_mapping = {}
    
    # Keywords to identify sheet types
    eoc_keywords = ["concession", "eoc", "conc", "concesión", "concesion"]
    eol_keywords = ["life", "eol", "vida", "útil", "util"]
    
    # Check each file's sheets
    for file_info in exploration_results["file_details"]:
        for sheet in file_info["sheets"]:
            sheet_lower = sheet.lower()
            
            # Check if sheet matches EOC pattern
            if any(keyword in sheet_lower.replace(" ", "") for keyword in eoc_keywords):
                sheet_mapping[sheet] = "end_of_concession"
            
            # Check if sheet matches EOL pattern
            elif any(keyword in sheet_lower.replace(" ", "") for keyword in eol_keywords):
                sheet_mapping[sheet] = "end_of_life"
            
            # Unknown type
            else:
                sheet_mapping[sheet] = "unknown"
    
    return sheet_mapping

def create_fixed_manifest(directory: str, output_path: str) -> None:
    """
    Create a fixed manifest.csv file based on the actual sheets in the Excel files.
    
    Args:
        directory: Directory containing the Excel files
        output_path: Path to save the fixed manifest.csv
    """
    directory = Path(directory)
    
    # Explore Excel files
    results = explore_excel_files(directory)
    
    # Build sheet type mapping
    sheet_mapping = build_sheet_mapping(results)
    
    # Create manifest data
    manifest_data = []
    
    for file_info in results["file_details"]:
        filename = file_info["filename"]
        year = file_info.get("year")
        engine = file_info.get("engine_used", "openpyxl")
        
        for sheet in file_info["sheets"]:
            scope = sheet_mapping.get(sheet, "unknown")
            
            manifest_data.append({
                "filename": filename,
                "year": year,
                "sheet_name": sheet,
                "scope": scope,
                "engine": engine
            })
    
    # Create DataFrame and save to CSV
    manifest_df = pd.DataFrame(manifest_data)
    manifest_df.to_csv(output_path, index=False)
    
    logger.info(f"Created fixed manifest with {len(manifest_df)} entries at {output_path}")
    print(f"Fixed manifest saved to {output_path}")

def preview_excel_content(file_path: str, num_rows: int = 5) -> Dict:
    """
    Preview the content of all sheets in an Excel file.
    
    Args:
        file_path: Path to the Excel file
        num_rows: Number of rows to preview from each sheet
    
    Returns:
        Dict: Dictionary with preview data for each sheet
    """
    file_path = Path(file_path)
    
    # Try both engines
    for engine in ['openpyxl', 'xlrd']:
        try:
            excel = pd.ExcelFile(file_path, engine=engine)
            break
        except Exception:
            continue
    else:
        return {"error": f"Could not open {file_path.name} with any engine"}
    
    preview_data = {}
    
    for sheet in excel.sheet_names:
        try:
            # Try first with no header
            df_no_header = pd.read_excel(file_path, sheet_name=sheet, header=None, nrows=num_rows, engine=engine)
            preview_data[f"{sheet}_raw"] = df_no_header.to_dict(orient='records')
            
            # Try to identify header row (simple approach - look at row 2, 3, 4)
            for header_row in [0, 1, 2, 3, 4]:
                if header_row < len(df_no_header):
                    df_with_header = pd.read_excel(
                        file_path, sheet_name=sheet, 
                        header=header_row, nrows=num_rows, engine=engine
                    )
                    preview_data[f"{sheet}_header{header_row}"] = df_with_header.to_dict(orient='records')
        except Exception as e:
            preview_data[f"{sheet}_error"] = str(e)
    
    return preview_data


analysis = analyze_sheet_differences("ArgGIS/data/raw/reserves")    
print(analysis["potential_sheet_mappings"])
